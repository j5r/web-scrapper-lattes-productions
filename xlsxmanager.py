from openpyxl import Workbook as WB




class Excel:
    def __init__(self):
        self.__nome = None
        self.__wb = WB()

        self.__wb.create_sheet("dados")
        self.__wb.create_sheet("lattes")
        self.__wb.create_sheet("orcid")

        self.__DADOS =  self.__wb.get_sheet_by_name("dados")
        self.__LATTES = self.__wb.get_sheet_by_name("lattes")
        self.__ORCID =  self.__wb.get_sheet_by_name("orcid")

        self.__DADOS.contador = 0
        self.__LATTES.contador = 0
        self.__ORCID.contador = 0

        try:
            self.__wb.remove(self.__wb.get_sheet_by_name("Sheet"))
        except Exception as e:
            print(e)

        # cabeçalhos
        self.__DADOS.append(["", "LATTES", "",
                "PUBLICADO ÚLTIMOS 5 ANOS *",
                "INDICADOR (JCR / QUALIS)",
                "COAUTORES DE INSTITUIÇÃO ESTRANGEIRA"])
        self.__DADOS.append(["","QTD Periódicos","ARTIGOS"])
        self.__LATTES.append(["","ANO","ARTIGO","DOI","JCR"])
        self.__ORCID.append(["","ANO","ARTIGO","DOI"])


    def lattes(self,**kw):
        kw.setdefault("ano","")
        kw.setdefault("artigo","")
        kw.setdefault("doi","")
        kw.setdefault("jcr","")

        kw = Excel.__treatNones(kw)
        try:
            self.__LATTES.contador += 1
            self.__LATTES.append([str(self.__LATTES.contador),
                            str(kw["ano"]).strip(),
                            str(kw["artigo"]).strip(),
                            str(kw["doi"]).strip(),
                            str(kw["jcr"]).strip()])
        except Exception as e:
            print(e)
            self.__LATTES.contador -= 1


    def orcid(self,**kw):
        kw.setdefault("ano","")
        kw.setdefault("artigo","")
        kw.setdefault("doi","")
        kw.setdefault("jcr","")

        kw = Excel.__treatNones(kw)
        try:
            self.__ORCID.contador += 1
            self.__ORCID.append([str(self.__ORCID.contador),
                        str(kw["ano"]).strip(),
                        str(kw["artigo"]).strip(),
                        str(kw["doi"]).strip()])
        except Exception as e:
            print(e)
            self.__ORCID.contador -= 1


    def dados(self,**kw):
        kw.setdefault("ano","")
        kw.setdefault("artigo","")
        kw.setdefault("jcr","")

        kw = Excel.__treatNones(kw)
        try:
            self.__DADOS.contador += 1
            self.__DADOS.append([str(self.__DADOS.contador),
                        str(self.__DADOS.contador),
                        str(kw["artigo"]).strip(),
                        str(kw["ano"]).strip(),
                        str(kw["jcr"]).strip()])
        except Exception as e:
            print(e)
            self.__DADOS.contador -= 1


    def save(self,nome=None):
        if self.__nome != None:
            self.__wb.save(self.__nome)
        elif nome != None:
            if not nome.endswith(".xlsx"):
                nome += ".xlsx"
            self.__nome = nome
            self.__wb.save(self.__nome)
        else:
            print("Dê um nome para o arquivo pelo menos uma vez.")

    def __treatNones(kw):
        for key in kw.keys():
            if kw[key] is None:
                kw[key] = "---"
            if kw[key] == "None":
                kw[key] = "---"
        return kw
