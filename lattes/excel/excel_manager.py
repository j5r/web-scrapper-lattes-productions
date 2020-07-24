from openpyxl import Workbook as WB
from lattes.mglobal.configs import JSON





def artigos_completos(arquivo, itens, wb):
    # nome das colunas da planilha excel gerada
    colunas_full_articles = JSON["excel"]["full_articles_columns"]


    full_articles = JSON["excel"]["full_articles_sheet_name"]
    wb.create_sheet(full_articles)
    FULL_ARTICLES = wb.get_sheet_by_name(full_articles)

    try: #tentando excluir a folha padrão "Sheet"
        wb.remove(wb.get_sheet_by_name("Sheet"))
    except Exception as e:
        print("Warning: ",e)

    FULL_ARTICLES.append(colunas_full_articles)


    for i in range(len(itens)):
        linha = [
            i+1,
            itens[i].ano,
            itens[i].jcr,
            itens[i].doi,
            itens[i].cite,
        ]
        FULL_ARTICLES.append(linha)



def trabalhos_completos(arquivo, itens, wb):
    # nome das colunas da planilha excel gerada
    colunas_full_works = JSON["excel"]["full_works_columns"]

    full_works = JSON["excel"]["full_works_sheet_name"]
    wb.create_sheet(full_works)
    FULL_WORKS = wb.get_sheet_by_name(full_works)

    FULL_WORKS.append(colunas_full_works)

    for i in range(len(itens)):
        linha = [
            i+1,
            itens[i].ano,
            itens[i].jcr,
            itens[i].doi,
            itens[i].cite,
        ]
        FULL_WORKS.append(linha)



def participacao_bancas_trabalhos_conclusao(arquivo,itens,wb):
    colunas_bancas = JSON["excel"]["exam_commissions_columns"]


    sheet_name = JSON["excel"]["exam_commissions_sheet_name"]
    wb.create_sheet(sheet_name)
    BANCAS = wb.get_sheet_by_name(sheet_name)

    BANCAS.append(colunas_bancas)

    ### append
    for i in range(len(itens)):
        linha = [i+1]
        linha.extend(itens[i])
        BANCAS.append(linha)



def membro_editorial(arquivo, itens, wb):
    colunas_membro = JSON["excel"]["editorial_member_columns"]
    sheet_name = JSON["excel"]["editorial_member_sheet_name"]
    wb.create_sheet(sheet_name)
    EDITORIAL = wb.get_sheet_by_name(sheet_name)

    EDITORIAL.append(colunas_membro)
    for i in range(len(itens)):
        linha = [i+1]
        linha.extend(itens[i])
        EDITORIAL.append(linha)


def revisor_de_periodico(arquivo, itens, wb):
    colunas_membro = JSON["excel"]["journal_reviewer_columns"]
    sheet_name = JSON["excel"]["journal_reviewer_sheet_name"]
    wb.create_sheet(sheet_name)
    REVIEWER = wb.get_sheet_by_name(sheet_name)

    REVIEWER.append(colunas_membro)
    for i in range(len(itens)):
        linha = [i+1]
        linha.extend(itens[i])
        REVIEWER.append(linha)


def projetos_pesquisa(arquivo, itens, wb):
    colunas_membro = JSON["excel"]["researcher_project_columns"]
    sheet_name = JSON["excel"]["researcher_project_sheet_name"]
    wb.create_sheet(sheet_name)
    PROJECTS = wb.get_sheet_by_name(sheet_name)

    PROJECTS.append(colunas_membro)
    for i in range(len(itens)):
        linha = [i+1]
        linha.extend(itens[i])
        PROJECTS.append(linha)
